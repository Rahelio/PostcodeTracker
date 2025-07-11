import logging
from datetime import datetime, timedelta
from functools import wraps
from flask import request, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
import jwt
from app import app
from database import db
from models import Journey, User
from postcode_service import PostcodeService

logger = logging.getLogger(__name__)

# Base path prefix for all API routes (matches Nginx alias)
API_PREFIX = '/LocationApp/api'

# JWT Token Management
def create_token(user_id: int) -> str:
    """Create a JWT token for the user."""
    payload = {
        'user_id': user_id,
        'exp': datetime.utcnow() + app.config['JWT_EXPIRATION_DELTA']
    }
    return jwt.encode(payload, app.config['JWT_SECRET_KEY'], algorithm='HS256')

def verify_token(token: str) -> int:
    """Verify JWT token and return user ID."""
    try:
        payload = jwt.decode(token, app.config['JWT_SECRET_KEY'], algorithms=['HS256'])
        logger.info(f"Token verified successfully for user_id: {payload['user_id']}")
        return payload['user_id']
    except jwt.ExpiredSignatureError:
        logger.warning("Token expired")
        return None
    except jwt.InvalidTokenError as e:
        logger.warning(f"Invalid token: {e}")
        return None

def get_current_user():
    """Get current user from JWT token (optional authentication)."""
    auth_header = request.headers.get('Authorization')
    if not auth_header:
        return None
    
    try:
        token = auth_header.split(' ')[1]  # Remove 'Bearer ' prefix
        user_id = verify_token(token)
        if user_id:
            return User.query.get(user_id)
    except (IndexError, AttributeError):
        pass
    
    return None

def require_auth(f):
    """Decorator to require authentication for endpoints."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        current_user = get_current_user()
        if not current_user:
            return jsonify({
                'success': False, 
                'message': 'Authentication required. Please log in.'
            }), 401
        return f(current_user, *args, **kwargs)
    return decorated_function

# API Routes
@app.route(f'{API_PREFIX}/health', methods=['GET'])
def health_check():
    """Health check endpoint."""
    return jsonify({
        'success': True,
        'message': 'PostcodeTracker API is running',
        'timestamp': datetime.utcnow().isoformat()
    })

@app.route(f'{API_PREFIX}/debug/user-count', methods=['GET'])
def debug_user_count():
    """Debug endpoint to check user count."""
    try:
        user_count = User.query.count()
        return jsonify({
            'success': True,
            'user_count': user_count,
            'timestamp': datetime.utcnow().isoformat()
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route(f'{API_PREFIX}/debug/active-journeys', methods=['GET'])
def debug_active_journeys():
    """Debug endpoint to check active journeys."""
    try:
        active_journeys = Journey.query.filter(Journey.end_time.is_(None)).all()
        return jsonify({
            'success': True,
            'active_journeys': [
                {
                    'id': j.id,
                    'user_id': j.user_id,
                    'username': j.user.username if j.user else 'Unknown',
                    'start_postcode': j.start_postcode,
                    'start_time': j.start_time.isoformat() if j.start_time else None
                } for j in active_journeys
            ],
            'count': len(active_journeys),
            'timestamp': datetime.utcnow().isoformat()
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route(f'{API_PREFIX}/debug/clear-active-journeys', methods=['POST'])
def debug_clear_active_journeys():
    """Debug endpoint to clear all active journeys."""
    try:
        active_journeys = Journey.query.filter(Journey.end_time.is_(None)).all()
        count = len(active_journeys)
        
        for journey in active_journeys:
            # Mark as ended with current time
            journey.end_time = datetime.utcnow()
            journey.end_postcode = journey.start_postcode  # Use start postcode as end
            journey.distance_miles = 0.0  # Zero distance for debug cleanup
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Cleared {count} active journeys',
            'cleared_count': count,
            'timestamp': datetime.utcnow().isoformat()
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route(f'{API_PREFIX}/auth/register', methods=['POST'])
def register():
    """Register a new user."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No data provided'}), 400
        
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        # Validation
        if not username or not password:
            return jsonify({'success': False, 'message': 'Username and password are required'}), 400
        
        if len(username) < 3:
            return jsonify({'success': False, 'message': 'Username must be at least 3 characters long'}), 400
        
        if len(password) < 6:
            return jsonify({'success': False, 'message': 'Password must be at least 6 characters long'}), 400
        
        # Check if user exists
        if User.query.filter_by(username=username).first():
            return jsonify({'success': False, 'message': 'Username already exists'}), 409
        
        # Create user
        password_hash = generate_password_hash(password)
        user = User(username=username, password_hash=password_hash)
        db.session.add(user)
        db.session.commit()
        
        # Create token
        token = create_token(user.id)
        
        logger.info(f"User {username} registered successfully")
        
        return jsonify({
            'success': True,
            'message': 'User registered successfully',
            'token': token,
            'user': user.to_dict()
        }), 201
        
    except Exception as e:
        logger.error(f"Registration error: {e}")
        db.session.rollback()
        return jsonify({'success': False, 'message': 'Registration failed'}), 500

@app.route(f'{API_PREFIX}/auth/login', methods=['POST'])
def login():
    """Login user."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No data provided'}), 400
        
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        if not username or not password:
            return jsonify({'success': False, 'message': 'Username and password are required'}), 400
        
        # Find user
        user = User.query.filter_by(username=username).first()
        if not user or not check_password_hash(user.password_hash, password):
            return jsonify({'success': False, 'message': 'Invalid username or password'}), 401
        
        # Create token
        token = create_token(user.id)
        
        response_data = {
            'success': True,
            'message': 'Login successful',
            'token': token,
            'user': user.to_dict()
        }
        
        logger.info(f"User {username} logged in successfully")
        logger.info(f"Login response: {response_data}")
        
        return jsonify(response_data)
        
    except Exception as e:
        logger.error(f"Login error: {e}")
        return jsonify({'success': False, 'message': 'Login failed'}), 500

@app.route(f'{API_PREFIX}/auth/profile', methods=['GET'])
@require_auth
def get_profile(current_user):
    """Get user profile information."""
    try:
        # Get user's journey statistics
        total_journeys = Journey.query.filter_by(user_id=current_user.id).count()
        completed_journeys = Journey.query.filter_by(user_id=current_user.id).filter(Journey.end_time.isnot(None)).count()
        total_distance = db.session.query(db.func.sum(Journey.distance_miles)).filter_by(user_id=current_user.id).scalar() or 0
        
        profile_data = current_user.to_dict()
        profile_data.update({
            'total_journeys': total_journeys,
            'completed_journeys': completed_journeys,
            'total_distance_miles': round(total_distance, 2)
        })
        
        return jsonify({
            'success': True,
            'data': profile_data
        })
        
    except Exception as e:
        logger.error(f"Error getting profile for user {current_user.username}: {e}")
        return jsonify({'success': False, 'message': 'Failed to get profile'}), 500

@app.route(f'{API_PREFIX}/auth/logout', methods=['POST'])
@require_auth
def logout(current_user):
    """Logout user (client should delete token)."""
    logger.info(f"User {current_user.username} logged out")
    return jsonify({
        'success': True,
        'message': 'Logged out successfully'
    })

@app.route(f'{API_PREFIX}/journey/start', methods=['POST'])
@require_auth
def start_journey(current_user):
    """Start a new journey using GPS coordinates. Requires authentication."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No data provided'}), 400
        
        latitude = data.get('latitude')
        longitude = data.get('longitude')
        client_name = data.get('client_name', '').strip()
        recharge_to_client = data.get('recharge_to_client', False)
        description = data.get('description', '').strip()
        
        if latitude is None or longitude is None:
            return jsonify({'success': False, 'message': 'Latitude and longitude are required'}), 400
        
        # Validate required fields
        if not client_name:
            return jsonify({'success': False, 'message': 'Client name is required'}), 400
        
        if not description:
            return jsonify({'success': False, 'message': 'Description is required'}), 400
        
        # Validate coordinates
        try:
            lat = float(latitude)
            lon = float(longitude)
            if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
                raise ValueError("Invalid coordinate range")
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'Invalid coordinates'}), 400
        
        # Check for active journey for this user
        active_journey = Journey.query.filter_by(user_id=current_user.id, end_time=None).first()
        if active_journey:
            return jsonify({
                'success': False,
                'message': 'You already have an active journey'
            }), 400
        
        # Get postcode from coordinates - timeout is handled by PostcodeService
        try:
            start_postcode = PostcodeService.get_postcode_from_coordinates(lat, lon)
            
            if not start_postcode:
                return jsonify({
                    'success': False, 
                    'message': f'Could not determine UK postcode for coordinates ({lat}, {lon}). This app only works within the UK. Please ensure you are in the UK and have a good GPS signal.'
                }), 400
                
        except TimeoutError:
            logger.error(f"Postcode lookup timed out for coordinates ({lat}, {lon})")
            return jsonify({
                'success': False, 
                'message': 'Postcode lookup timed out. Please try again or move to a different location.'
            }), 408
        except Exception as e:
            logger.error(f"Error during postcode lookup: {e}")
            return jsonify({
                'success': False, 
                'message': f'Could not determine UK postcode for coordinates ({lat}, {lon}). This app only works within the UK. Please ensure you are in the UK and have a good GPS signal.'
            }), 400
        
        # Create new journey for this user
        journey = Journey(
            start_postcode=start_postcode,
            start_latitude=lat,
            start_longitude=lon,
            user_id=current_user.id,
            client_name=client_name,
            recharge_to_client=recharge_to_client,
            description=description
        )
        db.session.add(journey)
        db.session.commit()
        
        logger.info(f"User {current_user.username} started journey {journey.id} with postcode {start_postcode}")
        
        return jsonify({
            'success': True,
            'message': 'Journey started successfully',
            'journey': journey.to_dict()
        }), 201
        
    except Exception as e:
        logger.error(f"Error starting journey: {e}")
        db.session.rollback()
        return jsonify({'success': False, 'message': 'Failed to start journey'}), 500

@app.route(f'{API_PREFIX}/journey/end', methods=['POST'])
@require_auth
def end_journey(current_user):
    """End the active journey. Requires authentication."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No data provided'}), 400
        
        latitude = data.get('latitude')
        longitude = data.get('longitude')
        
        if latitude is None or longitude is None:
            return jsonify({'success': False, 'message': 'Latitude and longitude are required'}), 400
        
        # Validate coordinates
        try:
            lat = float(latitude)
            lon = float(longitude)
            if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
                raise ValueError("Invalid coordinate range")
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'Invalid coordinates'}), 400
        
        # Find active journey for this user
        journey = Journey.query.filter_by(user_id=current_user.id, end_time=None).first()
        if not journey:
            return jsonify({'success': False, 'message': 'No active journey found'}), 404
        
        # Get end postcode from coordinates - timeout is handled by PostcodeService
        try:
            end_postcode = PostcodeService.get_postcode_from_coordinates(lat, lon)
            
            if not end_postcode:
                return jsonify({
                    'success': False, 
                    'message': f'Could not determine UK postcode for coordinates ({lat}, {lon}). This app only works within the UK. Please ensure you are in the UK and have a good GPS signal.'
                }), 400
                
        except TimeoutError:
            logger.error(f"Postcode lookup timed out for coordinates ({lat}, {lon})")
            return jsonify({
                'success': False, 
                'message': 'Postcode lookup timed out. Please try again or move to a different location.'
            }), 408
        except Exception as e:
            logger.error(f"Error during postcode lookup: {e}")
            return jsonify({
                'success': False, 
                'message': f'Could not determine UK postcode for coordinates ({lat}, {lon}). This app only works within the UK. Please ensure you are in the UK and have a good GPS signal.'
            }), 400
        
        # Calculate distance
        distance = PostcodeService.calculate_distance(journey.start_postcode, end_postcode)
        if distance is None:
            logger.warning(f"Could not calculate distance between {journey.start_postcode} and {end_postcode}")
            # Continue anyway - distance calculation failure shouldn't stop journey completion
        
        # Update journey
        journey.end_postcode = end_postcode
        journey.end_latitude = lat
        journey.end_longitude = lon
        journey.end_time = datetime.utcnow()
        journey.distance_miles = distance
        
        db.session.commit()
        
        logger.info(f"Journey {journey.id} ended with postcode {end_postcode}, distance: {distance}")
        
        return jsonify({
            'success': True,
            'message': 'Journey ended successfully',
            'journey': journey.to_dict()
        })
        
    except Exception as e:
        logger.error(f"Error ending journey: {e}")
        db.session.rollback()
        return jsonify({'success': False, 'message': 'Failed to end journey'}), 500

@app.route(f'{API_PREFIX}/journey/manual', methods=['POST'])
@require_auth
def create_manual_journey(current_user):
    """Create a manual journey using postcodes. Requires authentication."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No data provided'}), 400
        
        start_postcode = data.get('start_postcode', '').strip().upper()
        end_postcode = data.get('end_postcode', '').strip().upper()
        
        # Get client information fields
        client_name = data.get('client_name')
        recharge_to_client = data.get('recharge_to_client')
        description = data.get('description')
        
        # Get journey date (optional)
        journey_date_str = data.get('journey_date')
        journey_date = datetime.utcnow()  # Default to current time
        
        if journey_date_str:
            try:
                # Parse ISO 8601 date string
                journey_date = datetime.fromisoformat(journey_date_str.replace('Z', '+00:00'))
            except (ValueError, AttributeError) as e:
                logger.warning(f"Invalid journey_date format: {journey_date_str}, using current time. Error: {e}")
                journey_date = datetime.utcnow()
        
        if not start_postcode or not end_postcode:
            return jsonify({'success': False, 'message': 'Start and end postcodes are required'}), 400
        
        # Validate postcodes (basic UK postcode pattern)
        import re
        postcode_pattern = r'^[A-Z]{1,2}[0-9][A-Z0-9]? ?[0-9][A-Z]{2}$'
        
        if not re.match(postcode_pattern, start_postcode) or not re.match(postcode_pattern, end_postcode):
            return jsonify({'success': False, 'message': 'Invalid postcode format'}), 400
        
        if start_postcode == end_postcode:
            return jsonify({'success': False, 'message': 'Start and end postcodes cannot be the same'}), 400
        
        # Calculate distance between postcodes
        distance = PostcodeService.calculate_distance(start_postcode, end_postcode)
        if distance is None:
            logger.warning(f"Could not calculate distance between {start_postcode} and {end_postcode}")
            # Continue anyway - we'll store the journey without distance
        
        # Create manual journey (already completed)
        journey = Journey(
            start_postcode=start_postcode,
            end_postcode=end_postcode,
            start_time=journey_date,
            end_time=journey_date,  # Manual journeys are immediately completed
            distance_miles=distance,
            user_id=current_user.id,
            client_name=client_name,
            recharge_to_client=recharge_to_client,
            description=description,
            # No coordinates for manual journeys
            start_latitude=None,
            start_longitude=None,
            end_latitude=None,
            end_longitude=None
        )
        
        db.session.add(journey)
        db.session.commit()
        
        logger.info(f"User {current_user.username} created manual journey {journey.id}: {start_postcode} to {end_postcode}, date: {journey_date}, distance: {distance}")
        
        return jsonify({
            'success': True,
            'message': 'Manual journey created successfully',
            'journey': journey.to_dict()
        }), 201
        
    except Exception as e:
        logger.error(f"Error creating manual journey: {e}")
        db.session.rollback()
        return jsonify({'success': False, 'message': 'Failed to create manual journey'}), 500

@app.route(f'{API_PREFIX}/journey/active', methods=['GET'])
@require_auth
def get_active_journey(current_user):
    """Get the current active journey for the authenticated user."""
    try:
        journey = Journey.query.filter_by(user_id=current_user.id, end_time=None).first()
        
        if journey:
            return jsonify({
                'success': True,
                'active': True,
                'journey': journey.to_dict()
            })
        else:
            return jsonify({
                'success': True,
                'active': False,
                'journey': None
            })
            
    except Exception as e:
        logger.error(f"Error getting active journey for user {current_user.username}: {e}")
        return jsonify({'success': False, 'message': 'Failed to get active journey'}), 500

@app.route(f'{API_PREFIX}/journeys', methods=['GET'])
@require_auth
def get_journeys(current_user):
    """Get all journeys for the authenticated user."""
    try:
        journeys = Journey.query.filter_by(user_id=current_user.id).order_by(Journey.start_time.desc()).all()
        
        logger.info(f"Retrieved {len(journeys)} journeys for user {current_user.username}")
        
        return jsonify({
            'success': True,
            'journeys': [journey.to_dict() for journey in journeys]
        })
        
    except Exception as e:
        logger.error(f"Error getting journeys for user {current_user.username}: {e}")
        return jsonify({'success': False, 'message': 'Failed to get journeys'}), 500

@app.route(f'{API_PREFIX}/postcodes', methods=['GET'])
def get_postcodes():
    """Legacy endpoint for postcodes - returns empty list since we removed postcode management."""
    return jsonify([])

@app.route(f'{API_PREFIX}/postcode/from-coordinates', methods=['GET'])
def get_postcode_from_coordinates():
    """Get UK postcode from coordinates."""
    try:
        latitude = request.args.get('latitude')
        longitude = request.args.get('longitude')
        
        if not latitude or not longitude:
            return jsonify({
                'success': False,
                'message': 'Latitude and longitude parameters are required'
            }), 400
        
        try:
            lat = float(latitude)
            lon = float(longitude)
        except (ValueError, TypeError):
            return jsonify({
                'success': False,
                'message': 'Invalid latitude or longitude format'
            }), 400
        
        postcode = PostcodeService.get_postcode_from_coordinates(lat, lon)
        
        if postcode:
            return jsonify({
                'success': True,
                'postcode': postcode
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Could not determine postcode for the given coordinates'
            }), 404
            
    except Exception as e:
        logger.error(f"Error getting postcode from coordinates: {e}")
        return jsonify({'success': False, 'message': 'Failed to get postcode'}), 500

@app.route(f'{API_PREFIX}/journeys/delete', methods=['POST'])
@require_auth
def delete_journeys(current_user):
    """Delete selected journeys for the current user."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No data provided'}), 400
        
        journey_ids = data.get('journey_ids', [])
        if not journey_ids or not isinstance(journey_ids, list):
            return jsonify({'success': False, 'message': 'Journey IDs are required'}), 400
        
        # Find and delete journeys belonging to this user
        deleted_count = 0
        for journey_id in journey_ids:
            journey = Journey.query.filter_by(id=journey_id, user_id=current_user.id).first()
            if journey:
                db.session.delete(journey)
                deleted_count += 1
        
        db.session.commit()
        
        logger.info(f"User {current_user.username} deleted {deleted_count} journey(s)")
        
        return jsonify({
            'success': True,
            'message': f'Successfully deleted {deleted_count} journey(s)',
            'deleted_count': deleted_count
        })
        
    except Exception as e:
        logger.error(f"Error deleting journeys for user {current_user.username}: {e}")
        db.session.rollback()
        return jsonify({'success': False, 'message': 'Failed to delete journeys'}), 500

@app.route(f'{API_PREFIX}/journeys/export/csv', methods=['GET'])
@require_auth
def export_journeys_csv(current_user):
    """Export completed journeys for the current user as a CSV file."""
    try:
        import csv
        from io import StringIO, BytesIO
        
        # Fetch journeys for user (include both completed and active journeys)
        journeys = Journey.query.filter_by(user_id=current_user.id).order_by(Journey.start_time.desc()).all()
        
        # Prepare CSV in memory
        csv_buffer = StringIO()
        writer = csv.writer(csv_buffer)
        
        # Header - new requested format
        writer.writerow([
            'Date', 'Postcode From', 'Postcode To', 'Client Name', 'Recharge to Client', 'Description', 'Total Miles'
        ])
        
        # Data rows
        for j in journeys:
            # Extract date component only (no time)
            start_date = ''
            
            try:
                if j.start_time:
                    start_date = j.start_time.strftime('%Y-%m-%d')
            except Exception as e:
                logger.warning(f"Error formatting date for journey {j.id}: {e}")
            
            # Handle Yes/No for recharge field
            recharge_text = 'Yes' if j.recharge_to_client else 'No' if j.recharge_to_client is not None else ''
            
            writer.writerow([
                start_date,                                          # Date (without time)
                j.start_postcode,                                   # Postcode From (Start Postcode)
                j.end_postcode or '',                               # Postcode To (End Postcode)
                j.client_name or '',                                # Client Name
                recharge_text,                                      # Recharge to Client (Yes/No)
                j.description or '',                                # Description
                f"{j.distance_miles:.2f}" if j.distance_miles else ''  # Total Miles
            ])
        
        # Convert to bytes
        csv_bytes = BytesIO(csv_buffer.getvalue().encode('utf-8'))
        
        # Build response
        csv_bytes.seek(0)
        return send_file(
            csv_bytes,
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'journeys_{current_user.username}.csv'
        )
    except Exception as e:
        logger.error(f"CSV export failed for user {current_user.username}: {e}")
        return jsonify({'success': False, 'message': 'Failed to export CSV'}), 500

@app.route(f'{API_PREFIX}/journeys/export/excel', methods=['GET'])
@require_auth
def export_journeys_excel(current_user):
    """Export completed journeys for the current user as an Excel (.xlsx) file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO
        
        # Fetch journeys for user (include both completed and active journeys)
        journeys = Journey.query.filter_by(user_id=current_user.id).order_by(Journey.start_time.desc()).all()
        
        # Create workbook and select active sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Journeys"
        
        # Header row with formatting
        headers = [
            'Date', 'Postcode From', 'Postcode To', 'Client Name', 'Recharge to Client', 'Description', 'Total Miles'
        ]
        
        # Add headers to first row
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            # Blue background with white text
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data rows
        for row_idx, journey in enumerate(journeys, 2):  # Start from row 2
            # Extract date component only (no time)
            start_date = ''
            try:
                if journey.start_time:
                    start_date = journey.start_time.strftime('%Y-%m-%d')
            except Exception as e:
                logger.warning(f"Error formatting date for journey {journey.id}: {e}")
            
            # Handle Yes/No for recharge field
            recharge_text = 'Yes' if journey.recharge_to_client else 'No' if journey.recharge_to_client is not None else ''
            
            # Fill row data
            row_data = [
                start_date,                                                    # Date (without time)
                journey.start_postcode,                                       # Postcode From (Start Postcode)
                journey.end_postcode or '',                                   # Postcode To (End Postcode) 
                journey.client_name or '',                                    # Client Name
                recharge_text,                                                # Recharge to Client (Yes/No)
                journey.description or '',                                    # Description
                float(journey.distance_miles) if journey.distance_miles else ''  # Total Miles
            ]
            
            for col, value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col, value=value)
        
        # Auto-adjust column widths for better appearance
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO buffer
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Build response
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'journeys_{current_user.username}.xlsx'
        )
        
    except Exception as e:
        logger.error(f"Excel export failed for user {current_user.username}: {e}")
        return jsonify({'success': False, 'message': 'Failed to export Excel file'}), 500

# --- JSON error handlers for unknown routes and methods ---
@app.errorhandler(404)
def handle_404(e):
    return jsonify({
        'success': False,
        'message': 'Not found',
        'error': 404
    }), 404

@app.errorhandler(405)
def handle_405(e):
    return jsonify({
        'success': False,
        'message': 'Method not allowed',
        'error': 405
    }), 405 